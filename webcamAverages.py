# -*- coding: utf-8 -*-
"""
Created on Sun Nov 05 16:44:56 2017

@author: Jordan
Processes coordinates recorded in webcam excel files
Takes median/mean of coordinates and saves to file
Currently sorts files with natural sorting using natsort function

"""

from __future__ import division
import numpy as np
import glob, os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from natsort import natsorted, ns # Natural sorting
import math

#distChestTable = 18 # From Nov 2 2017 testing
distChestTable = 346 * 10 / 505# distance from chest to marker, in cm. Based on a calibration picture. Nov 9 2017
scaleCameraTable = 73.5 / 5.0 # Calibration, pixels / cm, nov 9 2017

#path = '20171102/Data/'
#files = glob.glob(path+ '*/Webcam_distMeas.xlsx')

#%% Functions

def checkColumns(measurements):
	# Checks the values in each column. Ignores -1 and returns the median
	colCheck = []
	rowStats=[]
	for i in range(0,measurements.shape[1]): # Iterate accross columns
		col = measurements[:,i] # Take one column at a time
		for cell in col: # Loop through each row of column
			if cell !=-1 and cell !=0:	# If cell is not in error (IE not = -1)
				colCheck.append(cell) # Append value to a checked list
		if colCheck:
			rowStats.append(np.median(colCheck)) # Take the Median instead of the mean
		else:
			rowStats.append(-1) # If no real values received at all, insert a -1
		colCheck = []  # Empty the checked column list
	return rowStats

def loadDistances(filepath):
	# Load XLSX file and return the data array of distance values
	wb = load_workbook(singleFile)
#	print wb.get_sheet_names()
	selectSheet = wb['Sheet1']['B3':'I52'] # Col B:K for a 5-50cm experiment, B:I for 5-40cm
	measurements = np.array([[i.value for i in j] for j in selectSheet]) # write each element into an array
	return measurements

def loadCoords(filepath):
	# Load XLSX file and return coordinate values
	# Function needs customization for each dataload
	# Customziation for Nov 9/10 recordings. 8 Distances, 50 samples (50x16) arrray
	wb = load_workbook(singleFile)
#	print wb.get_sheet_names()
	selectSheet = wb['Sheet2']['B4':'Q53'] 
	P1 = np.array([[i.value if type(i.value) != type(None) else 0 for i in j] for j in selectSheet]) 
	selectSheet = wb['Sheet3']['B4':'Q53'] 
	P2 = np.array([[i.value if type(i.value) != type(None) else 0 for i in j] for j in selectSheet]) # write each element into an array
	selectSheet = wb['Sheet4']['B4':'Q53'] 
	P3 = np.array([[i.value if type(i.value) != type(None) else 0 for i in j] for j in selectSheet])  # write each element into an array
	return P1,P2,P3

def coordAggregate(coords):
	# Aggregate (Mean/Median) values for all trials
	coordTuple = []
	for trial in coords: # Go through each trial
		xtrial = []
		for measurement in trial: # Grab one coordinate set at a time
			y = np.ma.masked_where(measurement==0,measurement) # Mask and ignore zeros
			x = np.ma.median(y,axis=0).filled(0) # Apply mask
			M = len(x)
			xlist=[]
			for i in range(int(M/2)):
				xlist.append((x[2*i],x[2*i+1]))
			xtrial.append(xlist)
		coordTuple.append(xtrial)
	return coordTuple

def calcAngle(p0,p1,p2):
	# p0 should be the anchor point. p1,p2 on the tag
	try:
		m1 = (p1[1] - p2[1]) / (p1[0] - p2[0]) # Slope between close points
		m2 = (p1[1] - p0[1]) / (p1[0] - p0[0]) # SLope to the anchor
		angle = math.atan((m1-m2) / (1+(m1*m2))) # angle in radians
		angle = angle * 180 # angle in degrees
		return angle
	except:
		return -1
	
def calcAngle2(p0,p1,p2):
	# p0 should be the anchor point. p1,p2 are on the tag
	try:
		m1 = (p1[1] - p2[1]) / (p1[0] - p2[0]) # Slope between close points
#		mid=(0,0) # Midpoint
		midx = (p1[0] + p2[0]) / 2 # Midpoint
		midy = (p1[1] + p2[1]) / 2 # Midpoint
		
		m2 = (midy - p0[1]) / (midx - p0[0]) # Slope from anchor to tag midpoint
		slope1 = math.atan(m1) * 180/3.14159 # Angle in degrees
		slope2 = math.atan(m2) * 180/3.14159
		print midx,midy, slope1,slope2
		angle =  slope2 - slope1# Angle between the two lines
		if angle < 0:
			angle = np.abs(angle)
		if angle > 0 and angle < 90:
			angle = 180 - angle
		
			
		return angle
	except:
		return -1
	
def dist(p0,p1):
	# Check distance
	return np.sqrt((p0[0] - p1[0])**2 + (p0[1] - p1[1])**2)

def pltSlopeField(slopes):
	# Plot Slope field with calculated Slopes
	for slope in slopes:
		x = slope[0][0]
		y = slope[0][1]
		m = slope[2]
		tol = 5
		domain = np.linspace(x-tol,x+tol,2)
		def fun(x1,y1):
			z = m*(domain-x1) + y1
			return z
		plt.plot(domain,fun(x,y),solid_capstyle='projecting',solid_joinstyle='bevel')
	plt.title('slope field')
	plt.grid(True)
	plt.show()
	return
	

#%% Load Files
#singleFile = files[0]
#wb = load_workbook(singleFile)
#print wb.get_sheet_names()
#selection = wb['Sheet1']['B3':'K52']
#measurements = np.array([[i.value for i in j] for j in wb['Sheet1']['B3':'K52']]) # write each element into an array

#%% Load all files, check values, and then export
# Nov2 files
path = '20171109/Data/'
files = glob.glob(path+ '*/Webcam_distMeas.xlsx')
files = natsorted(files)
stats = []
coords = []
singleFile = files[0]
for singleFile in files:
	print singleFile
	distances = loadDistances(singleFile)
	P1,P2,P3 = loadCoords(singleFile)
	coords.append([P1,P2,P3])
#	coords.append([checkColumns(P1),checkColumns(P2),checkColumns(P3)]) # This is the better way to do things in long run
	rowStats = checkColumns(distances)
	stats.append(rowStats)
statsArray = np.array(stats)
	
path = '20171109/Data/'
files = glob.glob(path+ '*/Webcam_distMeas.xlsx')
files = natsorted(files)
#stats = []
singleFile = files[0]
for singleFile in files:
	print singleFile
	distances = loadDistances(singleFile)
	P1,P2,P3 = loadCoords(singleFile)
	coords.append([P1,P2,P3])
#	coords.append([checkColumns(P1),checkColumns(P2),checkColumns(P3)]) # This is the better way to do things in long run
	rowStats = checkColumns(distances)
	stats.append(rowStats)
#statsArray = np.vstack((statsArray,stats))
statsArray = np.array(stats)
statsArraycm = statsArray / scaleCameraTable

#%% Process each trial data
# Process each file, mask errors, take median, summaize
# List hiearchy is trial > coordinate (1,2,3) > distance tuple
coordTuple = coordAggregate(coords) # Take the median of all the coords, return as a list

#%% Processing - 
distThresh = 60 # Pixel threshold distance, helping separate the two trackers on hand from anchor tracker
angles = []
negatives = []
for trial in coordTuple:
	angdist = []
#	print len(trial[0])
	for i in range(len(trial[0])):
		p1 = trial[1][i]
		p2 = trial[2][i]
		p0 = trial[0][i] # Should usually be our anchor point
		if dist(p1,p2) < distThresh:
			angle = calcAngle2(p0,p1,p2)
			angdist.append(angle)
			if angle < 0:
				negatives.append([p1,p2,p0])
		else:
			print 'Issue with assumed points'
			angdist.append(-1)
		
	angles.append(angdist)
	angdist = []

allangles = [angle for trial in angles for angle in trial if angle == angle] # Grab all the non NAN angles
plt.hist(allangles)
fig,ax = plt.subplots()
plt.ylim((0,1024))
plt.xlim((0,768))
for i in range(4):
	plt.scatter(*zip(*coordTuple[2*i][0]),marker='x',label='stationary') # Stationary marker
	plt.scatter(*zip(*coordTuple[2*i][1]),marker='^',label='tag1')
	plt.scatter(*zip(*coordTuple[2*i][2]),marker='o',label='tag2')
#ax.legend()
plt.show()

#%% Slope field
allangles = [angle for trial in angles for angle in trial] #this time including NaN
allcoords = []
slopes=[]
for trial in coordTuple:
	for i in range(8):
		allcoords.append([trial[0][i], trial[1][i], trial[2][i]])
for angle,coord in zip(allangles,allcoords):
	if angle > 0:
		p1 = coord[1]
		p2 = coord[2]
		slope = (p2[1] - p1[1]) / (p2[0] - p1[0])
		slopes.append([p1,p2,slope])
		
# Plot slope field
pltSlopeField(slopes)

		
#%% Combine and export	
statsArray = np.vstack((statsArray,np.array(stats)))
statsArray = np.transpose(statsArray)
statsArraycm = statsArray / scaleCameraTable

np.savetxt('investigation3_cm.csv',statsArraycm,delimiter=',')
np.savetxt('investigation3_angles.csv',allangles,delimiter=',')
